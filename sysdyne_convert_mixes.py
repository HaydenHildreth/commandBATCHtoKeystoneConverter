"""
convert_mix_listing_v2.py
author Hayden Hildreth
version 0.1.0
revision date 04/14/2026

Convert new vendor's mix design export (xlsx, one mix per row) into a nice
and readable format which can be imported into Keystone.

Input format:
  Each row is one mix design. Columns of interest:
    Code:          Mix design name
    MaterialName1 / MaterialAmount1 / MaterialUnit1  } repeating group
    MaterialName2 / MaterialAmount2 / MaterialUnit2  } up to 16 materials
    ...

Output:
  Excel Column A (col 0): Mix Design Name
  Excel Column B (col 1): Ingredient name
  Excel Column C (col 2): Unit (LB / OZ / etc.)
  Excel Column D (col 3): Amount

Usage:
  python convert_mix_listing_v2.py [input.xlsx] [output.xlsx] [plant_separator]

Defaults (these are used if script is ran without parameters at runtime):
  input  = new_vendor_export.xlsx
  output = MixListingEdit_converted_v2.xlsx
  plant_separator = ""
"""

import sys
import pandas as pd
import openpyxl

INPUT           = sys.argv[1] if len(sys.argv) > 1 else "new_vendor_export.xlsx"
OUTPUT          = sys.argv[2] if len(sys.argv) > 2 else "MixListingEdit_converted_v2.xlsx"
PLANT_SEPARATOR = sys.argv[3] if len(sys.argv) > 3 else ""

MAX_MATERIALS = 16  # Vendor supports up to 16 material slots per mix


def parse_mixes(path):
    """
    Read new vendor xlsx into a list of mix dicts:
      { 'name': str, 'ingredients': [(name, amount_str, unit), ...] }
    Skips ingredients where the name is empty or amount is 0.
    """
    df = pd.read_excel(path, dtype=str)

    mixes = []

    for _, row in df.iterrows():
        name = str(row.get('Code', '')).strip()
        if not name or name == 'nan':
            continue

        ingredients = []
        for n in range(1, MAX_MATERIALS + 1):
            mat_name   = row.get(f'MaterialName{n}',   None)
            mat_amount = row.get(f'MaterialAmount{n}', None)
            mat_unit   = row.get(f'MaterialUnit{n}',   None)

            if pd.isna(mat_name) or str(mat_name).strip() == '':
                continue

            mat_name = str(mat_name).strip()
            mat_unit = str(mat_unit).strip() if not pd.isna(mat_unit) else ''

            try:
                amount_f = float(mat_amount) if not pd.isna(mat_amount) else 0.0
            except (ValueError, TypeError):
                amount_f = 0.0

            # Skip zero-amount ingredients
            if amount_f == 0.0:
                continue

            amount_str = f"{amount_f:.3f}"
            ingredients.append((mat_name, amount_str, mat_unit))

        if ingredients:
            mixes.append({'name': name, 'ingredients': ingredients})

    return mixes


def write_output(mixes, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    out_row = 1  # openpyxl is 1-indexed

    for mix in mixes:
        name = mix['name']

        for (ingredient, amount, ing_unit) in mix['ingredients']:
            ws.cell(out_row, 1, (name + PLANT_SEPARATOR).upper())
            ws.cell(out_row, 2, (ingredient + PLANT_SEPARATOR).upper())
            ws.cell(out_row, 3, ing_unit.upper())
            ws.cell(out_row, 4, amount)
            out_row += 1

    wb.save(path)
    print(f"Written {out_row - 1} rows across {len(mixes)} mixes -> {path}")


if __name__ == "__main__":
    print(f"Parsing {INPUT} ...")
    mixes = parse_mixes(INPUT)
    print(f"Found {len(mixes)} mix designs.")
    if PLANT_SEPARATOR:
        print(f"Using plant separator: {repr(PLANT_SEPARATOR)}")
    write_output(mixes, OUTPUT)
